from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import JSONResponse, PlainTextResponse
import uvicorn
import os
import requests
import time
import json
from datetime import datetime
import sys
import xml.etree.ElementTree as ET
from io import BytesIO
from database import SalesDatabase
sys.path.insert(0, '/opt/tricycle-agent')
from WXBizMsgCrypt3 import WXBizMsgCrypt

app = FastAPI(title="電動三輪車智能體系統")

# 企業微信配置
CORP_ID = "wwaf0a27c133aa5659"
AGENT_ID = "1000002"
SECRET = "nA1J1J2gXKADsvhUAT18aIwg6l0o-dcl2ABFglUmFko"
TOKEN = "tricycle2026"
ENCODING_AES_KEY = "Aa1lRlBRltsnBFOoFer3jDboyNCxKaFco9JvG4uDG0p"

# Access Token 緩存
access_token_cache = {
    "token": None,
    "expires_at": 0
}

def get_access_token():
    """獲取企業微信 Access Token"""
    now = time.time()
    
    if access_token_cache["token"] and now < access_token_cache["expires_at"]:
        return access_token_cache["token"]
    
    url = f"https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={CORP_ID}&corpsecret={SECRET}"
    try:
        resp = requests.get(url, timeout=10)
        data = resp.json()
        
        if data.get("errcode") == 0:
            access_token_cache["token"] = data["access_token"]
            access_token_cache["expires_at"] = now + data["expires_in"] - 300
            return data["access_token"]
        else:
            print(f"獲取 access_token 失敗: {data}")
            return None
    except Exception as e:
        print(f"獲取 access_token 異常: {e}")
        return None

def download_wechat_media(media_id):
    """下載企業微信媒體文件"""
    try:
        access_token = get_access_token()
        if not access_token:
            return None, None
        
        url = f"https://qyapi.weixin.qq.com/cgi-bin/media/get?access_token={access_token}&media_id={media_id}"
        
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200:
            # 從 Content-Disposition 獲取文件名
            filename = "file"
            if 'Content-Disposition' in resp.headers:
                import re
                match = re.search(r'filename="?([^"]+)"?', resp.headers['Content-Disposition'])
                if match:
                    filename = match.group(1)
            
            return resp.content, filename
        else:
            print(f"下載文件失敗: {resp.text}")
            return None, None
    except Exception as e:
        print(f"下載文件異常: {e}")
        return None, None

def analyze_excel_v2(file_content, filename):
    """分析 Excel 文件並存入數據庫"""
    try:
        import openpyxl
        
        # 讀取 Excel
        wb = openpyxl.load_workbook(BytesIO(file_content))
        sheet = wb.active
        
        # 初始化數據庫
        db = SalesDatabase()
        
        # 解析數據
        data_list = []
        headers = []
        
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i == 1:
                headers = [str(cell).strip() if cell else f'col_{j}' for j, cell in enumerate(row)]
                continue
            
            if not any(row):
                continue
            
            row_data = {}
            for j, cell in enumerate(row):
                if j < len(headers):
                    row_data[headers[j]] = cell
            
            data_item = {
                'region': '',
                'customer': '',
                'product': '',
                'quantity': 0,
                'amount': 0.0,
                'date': ''
            }
            
            for key, value in row_data.items():
                key_lower = key.lower()
                
                if any(k in key_lower for k in ['区域', '地区', 'region', '省']):
                    data_item['region'] = str(value) if value else ''
                elif any(k in key_lower for k in ['客户', '客戶', 'customer', '姓名', '名称']):
                    data_item['customer'] = str(value) if value else ''
                elif any(k in key_lower for k in ['产品', '產品', 'product', '型号', '型號']):
                    data_item['product'] = str(value) if value else ''
                elif any(k in key_lower for k in ['数量', '數量', 'quantity', '台数', '台數']):
                    try:
                        data_item['quantity'] = int(float(value)) if value else 0
                    except:
                        data_item['quantity'] = 0
                elif any(k in key_lower for k in ['金额', '金額', 'amount', '价格', '價格']):
                    try:
                        data_item['amount'] = float(value) if value else 0.0
                    except:
                        data_item['amount'] = 0.0
                elif any(k in key_lower for k in ['日期', 'date', '时间', '時間']):
                    data_item['date'] = str(value) if value else ''
            
            data_list.append(data_item)
        
        # 清空舊數據並插入新數據
        db.clear_all()
        db.insert_sales_data(data_list)
        
        # 生成統計報告
        region_summary = db.get_region_summary()
        top_customers = db.get_top_customers(5)
        
        result = f"📊 Excel 數據已導入數據庫\n\n"
        result += f"文件名：{filename}\n"
        result += f"總記錄數：{len(data_list)} 條\n\n"
        
        result += f"【區域統計】\n"
        for r in region_summary[:5]:
            result += f"• {r['region']}：{r['total_quantity']} 台\n"
        
        result += f"\n【TOP 5 客戶】\n"
        for i, c in enumerate(top_customers, 1):
            result += f"{i}. {c['customer']} ({c['region']})：{c['total_quantity']} 台\n"
        
        result += f"\n✅ 現在可以在企業微信查詢真實數據！"
        
        return result
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"❌ 數據導入失敗：{str(e)}"

def create_text_reply(to_user, from_user, content):
    """創建文本回覆消息（被動回覆）"""
    timestamp = str(int(time.time()))
    nonce = str(int(time.time() * 1000))
    
    # 構建回覆 XML
    reply_xml = f"""<xml>
<ToUserName><![CDATA[{to_user}]]></ToUserName>
<FromUserName><![CDATA[{from_user}]]></FromUserName>
<CreateTime>{timestamp}</CreateTime>
<MsgType><![CDATA[text]]></MsgType>
<Content><![CDATA[{content}]]></Content>
</xml>"""
    
    # 加密回覆
    try:
        wxcpt = WXBizMsgCrypt(TOKEN, ENCODING_AES_KEY, CORP_ID)
        ret, encrypt_xml = wxcpt.EncryptMsg(reply_xml, nonce, timestamp)
        
        if ret == 0:
            return encrypt_xml
        else:
            print(f"加密失敗，錯誤碼: {ret}")
            return None
    except Exception as e:
        print(f"加密異常: {e}")
        import traceback
        traceback.print_exc()
        return None

def process_message(from_user, to_user, msg_type, content, media_id=None):
    """處理接收到的消息並返回回覆"""
    print(f"處理消息: from={from_user}, type={msg_type}, content={content}, media_id={media_id}")
    
    if msg_type == "text":
        if "你好" in content or "hello" in content.lower():
            reply = "你好！我係三輪車智能助手，有咩可以幫到你？\n\n試下發送：\n• 銷量\n• 任務\n• 幫助\n\n或者直接上傳 Excel 報表！"
        elif "銷量" in content or "销量" in content:
            # 智能查詢數據庫
            db = SalesDatabase()
            
            # 檢查是否有關鍵詞
            keywords = content.replace("銷量", "").replace("销量", "").strip()
            
            if keywords:
                # 搜尋特定關鍵詞
                results = db.search(keywords)
                if results:
                    reply = f"📊 查詢結果：{keywords}\n\n"
                    for r in results[:5]:
                        reply += f"• {r['region']} - {r['customer']}\n"
                        reply += f"  銷量：{r['total_quantity']} 台\n"
                else:
                    reply = f"❌ 未找到「{keywords}」相關數據"
            else:
                # 顯示整體統計
                region_summary = db.get_region_summary()
                top_customers = db.get_top_customers(3)
                
                if region_summary:
                    total = sum(r['total_quantity'] for r in region_summary)
                    reply = f"📊 銷量統計\n\n總銷量：{total} 台\n\n【區域分佈】\n"
                    for i, r in enumerate(region_summary[:5], 1):
                        emoji = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"][i-1] if i <= 5 else "•"
                        reply += f"{emoji} {r['region']}：{r['total_quantity']} 台\n"
                    
                    reply += f"\n【TOP 客戶】\n"
                    for i, c in enumerate(top_customers, 1):
                        reply += f"{i}. {c['customer']}：{c['total_quantity']} 台\n"
                    
                    reply += f"\n💡 試試發送「山東銷量」或「泰安銷量」查詢特定數據"
                else:
                    reply = "❌ 暫無數據，請先上傳 Excel 報表\n\n訪問：http://123.207.255.203/upload"
        elif "幫助" in content or "help" in content.lower():
            reply = "📋 功能列表\n\n1️⃣ 查詢銷量 - 發送「銷量」\n2️⃣ 上傳報表 - 直接發送 Excel 文件\n3️⃣ 任務管理 - 發送「任務」\n\n更多功能開發中..."
        elif "任務" in content:
            reply = "📋 任務管理\n\n【進行中】3 項\n• A型號生產計劃 (張三)\n• 採購原材料 (李四)\n• 品質檢查 (王五)\n\n【已完成】12 項\n【延誤】1 項 ⚠️"
        else:
            reply = f"收到你嘅消息：{content}\n\n試下發送：\n• 銷量\n• 任務\n• 幫助\n\n或者上傳 Excel 報表！"
        
        return create_text_reply(from_user, to_user, reply)
    
    elif msg_type == "file":
        # 處理文件上傳
        reply = "📄 正在處理文件，請稍候..."
        
        # 先發送處理中的消息
        processing_reply = create_text_reply(from_user, to_user, reply)
        
        # 下載並分析文件
        file_content, filename = download_wechat_media(media_id)
        
        if file_content:
            if filename.endswith(('.xlsx', '.xls')):
                # 分析 Excel
                result = analyze_excel_v2(file_content, filename)
            else:
                result = f"✅ 文件已接收：{filename}\n\n暫時只支持 Excel 文件分析。"
            
            return create_text_reply(from_user, to_user, result)
        else:
            return create_text_reply(from_user, to_user, "❌ 文件下載失敗，請重試")
    
    return None

@app.get("/")
async def root():
    return {
        "status": "running",
        "service": "電動三輪車智能體系統",
        "version": "2.1.0",
        "wechat": "已配置",
        "callback": "已啟用",
        "reply_mode": "被動回覆（無需IP白名單）",
        "features": ["文本回覆", "Excel分析"],
        "time": datetime.now().isoformat()
    }

@app.get("/health")
async def health():
    token = get_access_token()
    wechat_status = "connected" if token else "disconnected"
    
    return {
        "status": "healthy",
        "wechat": wechat_status,
        "callback": "ready",
        "reply_mode": "passive"
    }

@app.get("/api/wechat/callback")
async def wechat_callback_get(
    msg_signature: str = "",
    timestamp: str = "",
    nonce: str = "",
    echostr: str = ""
):
    """企業微信 URL 驗證"""
    print(f"收到驗證請求: signature={msg_signature}, timestamp={timestamp}, nonce={nonce}")
    
    try:
        wxcpt = WXBizMsgCrypt(TOKEN, ENCODING_AES_KEY, CORP_ID)
        ret, sEchoStr = wxcpt.VerifyURL(msg_signature, timestamp, nonce, echostr)
        
        if ret != 0:
            print(f"驗證失敗，錯誤碼: {ret}")
            return PlainTextResponse("驗證失敗")
        
        print(f"驗證成功: {sEchoStr}")
        return PlainTextResponse(sEchoStr.decode('utf-8'))
    except Exception as e:
        print(f"驗證異常: {e}")
        import traceback
        traceback.print_exc()
        return PlainTextResponse(f"驗證異常: {str(e)}")

@app.post("/api/wechat/callback")
async def wechat_callback_post(
    request: Request,
    msg_signature: str = "",
    timestamp: str = "",
    nonce: str = ""
):
    """企業微信消息接收（被動回覆模式）"""
    try:
        body = await request.body()
        print(f"收到企業微信消息: {body[:200]}")
        
        # 解密消息
        wxcpt = WXBizMsgCrypt(TOKEN, ENCODING_AES_KEY, CORP_ID)
        ret, xml_content = wxcpt.DecryptMsg(body, msg_signature, timestamp, nonce)
        
        if ret != 0:
            print(f"解密失敗，錯誤碼: {ret}")
            return PlainTextResponse("success")
        
        print(f"解密成功: {xml_content}")
        
        # 解析 XML
        root = ET.fromstring(xml_content)
        msg_type = root.find("MsgType").text
        from_user = root.find("FromUserName").text
        to_user = root.find("ToUserName").text
        
        if msg_type == "text":
            content = root.find("Content").text
            print(f"收到文本消息: from={from_user}, content={content}")
            
            reply_xml = process_message(from_user, to_user, msg_type, content)
            
            if reply_xml:
                print(f"返回被動回覆")
                return PlainTextResponse(reply_xml, media_type="application/xml")
        
        elif msg_type == "file":
            media_id = root.find("MediaId").text
            print(f"收到文件: from={from_user}, media_id={media_id}")
            
            reply_xml = process_message(from_user, to_user, msg_type, None, media_id=media_id)
            
            if reply_xml:
                print(f"返回被動回覆")
                return PlainTextResponse(reply_xml, media_type="application/xml")
        
        return PlainTextResponse("success")
    except Exception as e:
        print(f"處理企業微信消息異常: {e}")
        import traceback
        traceback.print_exc()
        return PlainTextResponse("success")

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """文件上傳接口（Web 版）"""
    try:
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        
        # 讀取文件內容
        file_content = await file.read()
        
        # 保存文件
        file_path = os.path.join(upload_dir, file.filename)
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        # 如果是 Excel，分析內容
        if file.filename.endswith(('.xlsx', '.xls')):
            result = analyze_excel(file_content, file.filename)
            return {
                "success": True,
                "filename": file.filename,
                "size": len(file_content),
                "analysis": result,
                "message": "文件上傳並分析成功"
            }
        else:
            return {
                "success": True,
                "filename": file.filename,
                "size": len(file_content),
                "message": "文件上傳成功（僅支持 Excel 分析）"
            }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e),
            "message": "上傳失敗"
        }

@app.get("/dashboard")
async def dashboard_page():
    """數據看板頁面"""
    from fastapi.responses import HTMLResponse
    with open('/opt/tricycle-agent/dashboard.html', 'r', encoding='utf-8') as f:
        return HTMLResponse(content=f.read())

@app.get("/api/dashboard/stats")
async def dashboard_stats():
    """數據看板統計接口"""
    from database import SalesDatabase
    db = SalesDatabase()
    
    # 獲取統計數據
    region_summary = db.get_region_summary()
    top_customers = db.get_top_customers(10)
    
    # 計算總計
    total_sales = sum(r['total_quantity'] for r in region_summary)
    total_amount = sum(r['total_amount'] for r in region_summary)
    
    return {
        "total_sales": total_sales,
        "region_count": len(region_summary),
        "customer_count": len(top_customers),
        "total_amount": total_amount,
        "regions": region_summary[:5],
        "customers": top_customers[:10]
    }


@app.get("/upload")
async def upload_page():
    """Web 上傳頁面"""
    from fastapi.responses import HTMLResponse
    with open('/opt/tricycle-agent/upload.html', 'r', encoding='utf-8') as f:
        return HTMLResponse(content=f.read())

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
def analyze_excel_v2(file_content, filename):
    """分析 Excel 文件並存入數據庫（優化版）"""
    try:
        import openpyxl
        from io import BytesIO
        from database import SalesDatabase
        
        # 讀取 Excel
        wb = openpyxl.load_workbook(BytesIO(file_content))
        sheet = wb.active
        
        # 初始化數據庫
        db = SalesDatabase()
        
        # 解析數據
        data_list = []
        headers = []
        
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i == 1:
                # 第一行作為表頭
                headers = [str(cell).strip() if cell else f'col_{j}' for j, cell in enumerate(row)]
                print(f"表頭: {headers}")
                continue
            
            # 跳過空行
            if not any(row):
                continue
            
            # 構建數據字典
            row_data = {}
            for j, cell in enumerate(row):
                if j < len(headers):
                    row_data[headers[j]] = cell
            
            # 智能識別字段
            data_item = {
                'region': '',
                'customer': '',
                'product': '',
                'quantity': 0,
                'amount': 0.0,
                'date': ''
            }
            
            # 嘗試匹配字段
            for key, value in row_data.items():
                key_lower = key.lower()
                
                # 區域：市级、区域、地区、省
                if '市级' in key or '市級' in key or '区域' in key_lower or '地区' in key_lower or 'region' in key_lower or '省' in key_lower or '城市' in key_lower:
                    data_item['region'] = str(value) if value else ''
                
                # 客戶：名称、客户、姓名
                elif '名称' in key or '名稱' in key or '客户' in key or '客戶' in key or 'customer' in key_lower or '姓名' in key or '商家' in key:
                    data_item['customer'] = str(value) if value else ''
                
                # 產品
                elif '产品' in key or '產品' in key or 'product' in key_lower or '型号' in key or '型號' in key:
                    data_item['product'] = str(value) if value else ''
                
                # 數量：销量、数量、台数
                elif '销量' in key or '銷量' in key or '数量' in key or '數量' in key or 'quantity' in key_lower or '台数' in key or '台數' in key or '汇总' in key or '匯總' in key:
                    try:
                        data_item['quantity'] = int(float(value)) if value else 0
                    except:
                        data_item['quantity'] = 0
                
                # 金額
                elif '金额' in key or '金額' in key or 'amount' in key_lower or '价格' in key or '價格' in key:
                    try:
                        data_item['amount'] = float(value) if value else 0.0
                    except:
                        data_item['amount'] = 0.0
                
                # 日期
                elif '日期' in key or 'date' in key_lower or '时间' in key or '時間' in key:
                    data_item['date'] = str(value) if value else ''
            
            data_list.append(data_item)
            print(f"行{i}: 區域={data_item['region']}, 客戶={data_item['customer'][:20]}, 數量={data_item['quantity']}")
        
        # 清空舊數據並插入新數據
        db.clear_all()
        db.insert_sales_data(data_list)
        
        # 生成統計報告
        region_summary = db.get_region_summary()
        top_customers = db.get_top_customers(5)
        
        result = f"📊 Excel 數據已導入數據庫\n\n"
        result += f"文件名：{filename}\n"
        result += f"總記錄數：{len(data_list)} 條\n\n"
        
        result += f"【區域統計】\n"
        for r in region_summary[:5]:
            result += f"• {r['region']}：{r['total_quantity']} 台\n"
        
        result += f"\n【TOP 5 客戶】\n"
        for i, c in enumerate(top_customers, 1):
            result += f"{i}. {c['customer'][:30]} ({c['region']})：{c['total_quantity']} 台\n"
        
        result += f"\n✅ 現在可以在企業微信查詢真實數據！"
        
        return result
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"❌ 數據導入失敗：{str(e)}"
